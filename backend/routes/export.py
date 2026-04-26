"""
Export route — generates comprehensive Excel reports.
"""
import os
from datetime import datetime
from fastapi import APIRouter, HTTPException
from fastapi.responses import FileResponse
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

from database.models import (
    get_file, get_financial_data, get_ratios, get_anomalies
)
from services.financial_engine import (
    build_yoy_table, calculate_revenue_growth, calculate_profit_margin,
    detect_trend
)

router = APIRouter()


def hex_to_rgb(hex_color):
    """Convert hex color to RGB tuple for openpyxl."""
    hex_color = hex_color.lstrip('#')
    return tuple(int(hex_color[i:i+2], 16) for i in (0, 2, 4))


def format_currency(value):
    """Format value as currency."""
    if value is None or value == 0:
        return 0
    return float(value)


@router.get("/export/{file_id}")
async def export_excel(file_id: int):
    """
    Export comprehensive financial report as Excel with 5 sheets:
    1. Cover - Summary statistics
    2. Raw Data - Financial data table
    3. Ratios - Financial metrics with benchmarks
    4. YoY Growth - Year-over-year analysis
    5. Anomaly Report - Detected anomalies
    """
    # Verify file exists
    file_record = get_file(file_id)
    if not file_record:
        raise HTTPException(status_code=404, detail="File not found")

    try:
        # Fetch data from database
        financial_data = get_financial_data(file_id)
        ratios = get_ratios(file_id)
        anomalies = get_anomalies(file_id)
        yoy_table = build_yoy_table(financial_data) if financial_data else []
        
        trend = detect_trend(financial_data) if financial_data else "No Data"
        latest_revenue = financial_data[-1].get("revenue", 0) if financial_data else 0
        latest_profit = financial_data[-1].get("profit", 0) if financial_data else 0

        # Create workbook
        wb = openpyxl.Workbook()
        
        # Define styles
        header_font = Font(bold=True, color="FFFFFF", size=11)
        header_fill = PatternFill(start_color="3B4A6B", end_color="3B4A6B", fill_type="solid")
        center_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
        
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )

        # ═══════════════════════════════════════════
        # SHEET 1: COVER
        # ═══════════════════════════════════════════
        ws_cover = wb.active
        ws_cover.title = "Cover"
        ws_cover.sheet_properties.tabColor = "4F46E5"

        # Title
        ws_cover['A1'] = file_record.get('filename', 'Financial Report')
        ws_cover['A1'].font = Font(bold=True, size=20, color="4F46E5")
        
        # Subtitle
        ws_cover['A2'] = "Financial Intelligence Report"
        ws_cover['A2'].font = Font(size=14, bold=True)
        
        # Date
        ws_cover['A3'] = f"Generated on: {datetime.now().strftime('%B %d, %Y')}"
        ws_cover['A3'].font = Font(size=10, color="555577")
        
        # Blank row
        # Row 5: Summary Statistics
        ws_cover['A5'] = "Summary Statistics"
        ws_cover['A5'].font = Font(bold=True, size=12)
        
        # Summary metrics
        metrics = [
            ("Total Years of Data", str(len(financial_data))),
            ("Latest Revenue", f"${latest_revenue:,.2f}"),
            ("Latest Profit", f"${latest_profit:,.2f}"),
            ("Revenue Growth (YoY)", f"{ratios.get('revenue_growth', 0):.2f}%"),
            ("Profit Margin", f"{ratios.get('profit_margin', 0):.2f}%"),
            ("EBITDA Margin", f"{ratios.get('ebitda_margin', 0):.2f}%"),
            ("Overall Trend", trend),
        ]
        
        for idx, (label, value) in enumerate(metrics, 6):
            ws_cover[f'A{idx}'] = label
            ws_cover[f'B{idx}'] = value
            ws_cover[f'A{idx}'].font = Font(size=10)
            ws_cover[f'B{idx}'].font = Font(size=10, bold=True)
        
        ws_cover.column_dimensions['A'].width = 30
        ws_cover.column_dimensions['B'].width = 20

        # ═══════════════════════════════════════════
        # SHEET 2: RAW DATA
        # ═══════════════════════════════════════════
        ws_data = wb.create_sheet("Raw Data")
        ws_data.sheet_properties.tabColor = "6366F1"
        
        # Headers
        headers = ["Year", "Revenue", "Profit", "Expenses", "EBITDA", "Total Assets",
                  "Total Liabilities", "Current Assets", "Current Liabilities", "Cash Flow"]
        for col, header in enumerate(headers, 1):
            cell = ws_data.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = center_align
            cell.border = thin_border
        
        # Data rows with alternating colors
        dark_fill_1 = PatternFill(start_color="1A1A3E", end_color="1A1A3E", fill_type="solid")
        dark_fill_2 = PatternFill(start_color="141432", end_color="141432", fill_type="solid")
        data_font = Font(size=10, color="FFFFFF")

        if not financial_data:
            no_data_cell = ws_data.cell(row=2, column=1, value="No data available")
            no_data_cell.font = Font(size=11, color="888888", italic=True)
        else:
            for row_idx, row_data in enumerate(financial_data, 2):
                fill = dark_fill_1 if row_idx % 2 == 0 else dark_fill_2

                values = [
                    row_data.get("year", ""),
                    format_currency(row_data.get("revenue", 0)),
                    format_currency(row_data.get("profit", 0)),
                    format_currency(row_data.get("expenses", 0)),
                    format_currency(row_data.get("ebitda", 0)),
                    format_currency(row_data.get("total_assets", 0)),
                    format_currency(row_data.get("total_liabilities", 0)),
                    format_currency(row_data.get("current_assets", 0)),
                    format_currency(row_data.get("current_liabilities", 0)),
                    format_currency(row_data.get("cash_flow", 0)),
                ]

                for col, value in enumerate(values, 1):
                    cell = ws_data.cell(row=row_idx, column=col, value=value)
                    cell.fill = fill
                    cell.font = data_font
                    cell.border = thin_border

                    if col > 1:  # Number columns
                        cell.number_format = '#,##0.00'
                        cell.alignment = Alignment(horizontal="right")
                    else:  # Year column
                        cell.alignment = center_align

        # Auto-fit columns (base on header length; data may vary)
        for col_idx, header in enumerate(headers, 1):
            max_length = max(len(str(header)), 15)
            if financial_data:
                for row_data in financial_data:
                    vals = list(row_data.values())
                    if col_idx - 1 < len(vals):
                        max_length = max(max_length, len(str(vals[col_idx - 1])))
            ws_data.column_dimensions[openpyxl.utils.get_column_letter(col_idx)].width = min(max_length + 2, 25)

        # ═══════════════════════════════════════════
        # SHEET 3: RATIOS
        # ═══════════════════════════════════════════
        ws_ratios = wb.create_sheet("Ratios")
        ws_ratios.sheet_properties.tabColor = "8B5CF6"
        
        # Headers
        ratio_headers = ["Metric", "Value", "Benchmark", "Status", "Description"]
        for col, header in enumerate(ratio_headers, 1):
            cell = ws_ratios.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = center_align
            cell.border = thin_border
        
        # Ratio data
        ratio_data = [
            ("Revenue Growth %", ratios.get("revenue_growth", 0), "> 10%", "Year-over-year revenue change"),
            ("Profit Margin %", ratios.get("profit_margin", 0), "> 15%", "Net profit as % of revenue"),
            ("EBITDA Margin %", ratios.get("ebitda_margin", 0), "> 20%", "Operating profit before D&A as % of revenue"),
            ("Debt/Equity Ratio", ratios.get("debt_equity", 0), "< 1.5", "Total debt relative to equity"),
            ("Current Ratio", ratios.get("current_ratio", 0), "> 1.5", "Ability to pay short-term obligations"),
            ("5-Year CAGR %", ratios.get("cagr_5yr", 0), "> 8%", "Compound annual growth rate of revenue"),
            ("Avg Profit Margin %", ratios.get("avg_profit_margin", 0), "> 12%", "Average margin across all years"),
        ]
        
        benchmarks = [10, 15, 20, 1.5, 1.5, 8, 12]
        # is_less_than: only Debt/Equity (index 3) is "lower is better"; all others are "higher is better"
        is_less_than = [False, False, False, True, False, False, False]
        
        for row_idx, (metric, value, benchmark, description) in enumerate(ratio_data, 2):
            # Metric name
            ws_ratios.cell(row=row_idx, column=1, value=metric).border = thin_border
            
            # Value
            value_cell = ws_ratios.cell(row=row_idx, column=2, value=value)
            value_cell.number_format = '0.00'
            value_cell.border = thin_border
            value_cell.alignment = Alignment(horizontal="right")
            
            # Benchmark
            ws_ratios.cell(row=row_idx, column=3, value=benchmark).border = thin_border
            
            # Status
            status_cell = ws_ratios.cell(row=row_idx, column=4)
            status_cell.border = thin_border
            status_cell.alignment = center_align
            
            if value == 0 or value is None:
                status_cell.value = "— N/A"
                status_cell.fill = PatternFill(start_color="F3F4F6", end_color="F3F4F6", fill_type="solid")
                status_cell.font = Font(color="555577")
            else:
                meets_benchmark = (value >= benchmarks[row_idx - 2]) if not is_less_than[row_idx - 2] else (value <= benchmarks[row_idx - 2])
                if meets_benchmark:
                    status_cell.value = "✓ Good"
                    status_cell.fill = PatternFill(start_color="D1FAE5", end_color="D1FAE5", fill_type="solid")
                    status_cell.font = Font(color="059669", bold=True)
                else:
                    status_cell.value = "⚠ Review"
                    status_cell.fill = PatternFill(start_color="FEE2E2", end_color="FEE2E2", fill_type="solid")
                    status_cell.font = Font(color="DC2626", bold=True)
            
            # Description
            ws_ratios.cell(row=row_idx, column=5, value=description).border = thin_border
        
        ws_ratios.column_dimensions['A'].width = 20
        ws_ratios.column_dimensions['B'].width = 12
        ws_ratios.column_dimensions['C'].width = 12
        ws_ratios.column_dimensions['D'].width = 12
        ws_ratios.column_dimensions['E'].width = 40

        # ═══════════════════════════════════════════
        # SHEET 4: YOY GROWTH
        # ═══════════════════════════════════════════
        ws_yoy = wb.create_sheet("YoY Growth")
        ws_yoy.sheet_properties.tabColor = "10B981"
        
        # Headers
        yoy_headers = ["Year", "Revenue", "Profit", "Revenue Growth %", "Profit Growth %", "Profit Margin %"]
        for col, header in enumerate(yoy_headers, 1):
            cell = ws_yoy.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = center_align
            cell.border = thin_border
        
        # YoY data
        if not yoy_table:
            no_yoy_cell = ws_yoy.cell(row=2, column=1, value="No data available")
            no_yoy_cell.font = Font(size=11, color="888888", italic=True)

        for row_idx, row_data in enumerate(yoy_table, 2):
            # Year
            ws_yoy.cell(row=row_idx, column=1, value=row_data.get("year", "")).border = thin_border
            
            # Revenue
            rev_cell = ws_yoy.cell(row=row_idx, column=2, value=format_currency(row_data.get("revenue", 0)))
            rev_cell.number_format = '#,##0.00'
            rev_cell.border = thin_border
            rev_cell.alignment = Alignment(horizontal="right")
            
            # Profit
            prof_cell = ws_yoy.cell(row=row_idx, column=3, value=format_currency(row_data.get("profit", 0)))
            prof_cell.number_format = '#,##0.00'
            prof_cell.border = thin_border
            prof_cell.alignment = Alignment(horizontal="right")
            
            # Revenue Growth %
            rev_growth = row_data.get("revenue_growth_pct")
            rev_growth_cell = ws_yoy.cell(row=row_idx, column=4, value=rev_growth if rev_growth is not None else None)
            rev_growth_cell.border = thin_border
            if rev_growth is None:
                rev_growth_cell.value = "—"
                rev_growth_cell.font = Font(color="555577")
                rev_growth_cell.alignment = center_align
            else:
                rev_growth_cell.number_format = '0.00'
                rev_growth_cell.alignment = Alignment(horizontal="right")
                if rev_growth > 0:
                    rev_growth_cell.fill = PatternFill(start_color="D1FAE5", end_color="D1FAE5", fill_type="solid")
                    rev_growth_cell.font = Font(color="065F46")
                elif rev_growth < 0:
                    rev_growth_cell.fill = PatternFill(start_color="FEE2E2", end_color="FEE2E2", fill_type="solid")
                    rev_growth_cell.font = Font(color="991B1B")
            
            # Profit Growth %
            prof_growth = row_data.get("profit_growth_pct")
            prof_growth_cell = ws_yoy.cell(row=row_idx, column=5, value=prof_growth if prof_growth is not None else None)
            prof_growth_cell.border = thin_border
            if prof_growth is None:
                prof_growth_cell.value = "—"
                prof_growth_cell.font = Font(color="555577")
                prof_growth_cell.alignment = center_align
            else:
                prof_growth_cell.number_format = '0.00'
                prof_growth_cell.alignment = Alignment(horizontal="right")
                if prof_growth > 0:
                    prof_growth_cell.fill = PatternFill(start_color="D1FAE5", end_color="D1FAE5", fill_type="solid")
                    prof_growth_cell.font = Font(color="065F46")
                elif prof_growth < 0:
                    prof_growth_cell.fill = PatternFill(start_color="FEE2E2", end_color="FEE2E2", fill_type="solid")
                    prof_growth_cell.font = Font(color="991B1B")
            
            # Profit Margin %
            margin = row_data.get("profit_margin_pct")
            margin_cell = ws_yoy.cell(row=row_idx, column=6, value=margin if margin is not None else None)
            margin_cell.number_format = '0.00'
            margin_cell.border = thin_border
            margin_cell.alignment = Alignment(horizontal="right")
        
        ws_yoy.column_dimensions['A'].width = 12
        ws_yoy.column_dimensions['B'].width = 15
        ws_yoy.column_dimensions['C'].width = 15
        ws_yoy.column_dimensions['D'].width = 16
        ws_yoy.column_dimensions['E'].width = 16
        ws_yoy.column_dimensions['F'].width = 16

        # ═══════════════════════════════════════════
        # SHEET 5: ANOMALY REPORT
        # ═══════════════════════════════════════════
        ws_anomaly = wb.create_sheet("Anomaly Report")
        ws_anomaly.sheet_properties.tabColor = "EF4444"
        
        if not anomalies:
            ws_anomaly['A1'] = "No anomalies detected. Financial data looks clean."
            ws_anomaly['A1'].font = Font(size=12, color="059669", bold=True)
        else:
            # Sort anomalies by severity then year
            severity_order = {"high": 0, "medium": 1, "low": 2}
            sorted_anomalies = sorted(
                anomalies,
                key=lambda x: (severity_order.get(x.get("severity", "medium"), 1), x.get("year", ""))
            )
            
            # Headers
            anomaly_headers = ["Year", "Field", "Actual Value", "Expected Value", "Deviation %", "Severity", "Description"]
            for col, header in enumerate(anomaly_headers, 1):
                cell = ws_anomaly.cell(row=1, column=col, value=header)
                cell.font = Font(bold=True, color="FFFFFF", size=11)
                cell.fill = PatternFill(start_color="7F1D1D", end_color="7F1D1D", fill_type="solid")
                cell.alignment = center_align
                cell.border = thin_border
            
            # Data rows
            for row_idx, anomaly in enumerate(sorted_anomalies, 2):
                severity = anomaly.get("severity", "medium").lower()
                
                # Year
                ws_anomaly.cell(row=row_idx, column=1, value=anomaly.get("year", "")).border = thin_border
                
                # Field
                field_val = str(anomaly.get("field", "")).replace("_", " ").title()
                ws_anomaly.cell(row=row_idx, column=2, value=field_val).border = thin_border
                
                # Actual Value
                actual_cell = ws_anomaly.cell(row=row_idx, column=3, value=format_currency(anomaly.get("value", 0)))
                actual_cell.number_format = '#,##0.00'
                actual_cell.border = thin_border
                
                # Expected Value
                expected_cell = ws_anomaly.cell(row=row_idx, column=4, value=format_currency(anomaly.get("expected_value", 0)))
                expected_cell.number_format = '#,##0.00'
                expected_cell.border = thin_border
                
                # Deviation %
                dev_cell = ws_anomaly.cell(row=row_idx, column=5, value=anomaly.get("deviation_pct", 0))
                dev_cell.number_format = '0.00'
                dev_cell.border = thin_border
                
                # Severity
                severity_cell = ws_anomaly.cell(row=row_idx, column=6, value=severity.upper())
                severity_cell.border = thin_border
                severity_cell.alignment = center_align
                severity_cell.font = Font(bold=True)
                
                if severity == "high":
                    severity_cell.fill = PatternFill(start_color="FEE2E2", end_color="FEE2E2", fill_type="solid")
                    severity_cell.font = Font(bold=True, color="DC2626")
                elif severity == "medium":
                    severity_cell.fill = PatternFill(start_color="FEF3C7", end_color="FEF3C7", fill_type="solid")
                    severity_cell.font = Font(bold=True, color="D97706")
                else:
                    severity_cell.fill = PatternFill(start_color="D1FAE5", end_color="D1FAE5", fill_type="solid")
                    severity_cell.font = Font(bold=True, color="059669")
                
                # Description
                desc_cell = ws_anomaly.cell(row=row_idx, column=7, value=anomaly.get("description", ""))
                desc_cell.border = thin_border
                desc_cell.alignment = Alignment(wrap_text=True)
            
            ws_anomaly.column_dimensions['A'].width = 10
            ws_anomaly.column_dimensions['B'].width = 15
            ws_anomaly.column_dimensions['C'].width = 15
            ws_anomaly.column_dimensions['D'].width = 15
            ws_anomaly.column_dimensions['E'].width = 12
            ws_anomaly.column_dimensions['F'].width = 12
            ws_anomaly.column_dimensions['G'].width = 40

        # Remove default sheet
        if "Sheet" in wb.sheetnames:
            del wb["Sheet"]

        # Save to file inside uploads/ directory (created on startup)
        os.makedirs("uploads", exist_ok=True)
        export_filename = f"uploads/export_{file_id}.xlsx"
        wb.save(export_filename)

        # Create response
        clean_filename = file_record.get('filename', 'report').replace('.pdf', '').replace(' ', '_')
        return FileResponse(
            path=export_filename,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            filename=f"FinancialReport_{clean_filename}.xlsx"
        )

    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Export failed: {str(e)}")
